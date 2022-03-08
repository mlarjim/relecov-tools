#!/usr/bin/env python
from itertools import islice
import logging
import rich.console
from openpyxl import Workbook
import openpyxl
import relecov_tools.utils

log = logging.getLogger(__name__)
stderr = rich.console.Console(
    stderr=True,
    style="dim",
    highlight=False,
    force_terminal=relecov_tools.utils.rich_force_colors(),
)


class RelecovMetadata:
    def __init__(
        self,
        metadata_file=None,
        additional_metadata_file=None,
        output_folder=None
    ):
        if metadata_file is None:
            self.metadata_file = relecov_tools.utils.prompt_path(
                msg="Select the excel file which contains metadata"
            )
        else:
            self.metadata_file = metadata_file
        if output_folder is None:
            self.output_folder = relecov_tools.utils.prompt_path(
                msg="Select the output folder"
            )
        else:
            self.output_folder = output_folder
        self.additional_metadata_file = additional_metadata_file


# read_metadata



    # Perform workflow details


def check_new_metadata(folder):
    """Check if there is a new metadata to be processed

    folder  Directory to be checked
    """
    pass


def fetch_metadata_file(folder, file_name):
    """Fetch the metadata file
    folder  Directory to fetch metadata file
    file_name   metadata file name
    """
    wb_file = openpyxl.load_workbook(file_name, data_only=True)
    ws_metadata_lab = wb_file["METADATA_LAB"]
    heading = []
    for cell in ws_metadata_lab[1]:
        heading.append(cell.value)



def validate_metadata_sample(row_sample):
    """Validate sample information"""


def add_extra_data(metadata_file, extra_data, result_metadata):
    """Add the additional information that must be included in final metadata
    metadata Origin metadata file
    extra_data  additional data to be included
    result_metadata    final metadata after adding the additional data
    """
    pass


def request_information(external_url, request):
    """Get information from external database server using Rest API

    external_url
    request
    """
    pass


def store_information(external_url, request, data):
    """Update information"""
    pass



def read_metadata_workflow(self):
    """
    Description :   Starts the read metada workflow
    """
    sample_list = []
    wb_file = openpyxl.load_workbook(arguments.inputFile, data_only=True)
    ws_metadata_lab = wb_file["METADATA_LAB"]
    heading = []
    for cell in ws_metadata_lab[1]:

        heading.append(cell.value)

    for i in range(len(heading)):
        if heading[i] in list(mapping_file.keys()):
            index = list(mapping_file).index(heading[i])
            heading[index] = list(mapping_file.values())[index]
    for row in islice(ws_metadata_lab.values, 1, ws_metadata_lab.max_row):
        sample_data_row = {}
        for idx in range(len(heading)):
            if "date" in heading[idx]:
                sample_data_row[heading[idx]] = row[idx].strftime("%d/%m/%Y")
            else:
                sample_data_row[heading[idx]] = row[idx]
        try:
            validate(instance=sample_data_row, schema=json_phage_plus_schema)
        except:
            print("Unsuccessful validation for sample ", sample_data_row["sample_name"])

            continue